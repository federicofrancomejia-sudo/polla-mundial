# -*- coding: utf-8 -*-
"""Migra las apuestas, claves y resultados del modelo Excel (POLLA INDIVIDUAL)
a la base de datos de la app web. Correr UNA vez antes de pasar al modelo web.

Uso local (SQLite):  py migrar_excel.py
Para la nube: define DATABASE_URL apuntando a Postgres antes de correr."""
import glob
import os

from openpyxl import load_workbook

import db
from fixtures import PARTIDOS

FOLDER = r"C:\Users\CO1015470664\OneDrive - Enel Spa\POLLA INDIVIDUAL"
TABLERO = os.path.join(FOLDER, "TABLERO.xlsx")


def main():
    nombres = []
    archivos = sorted(glob.glob(os.path.join(FOLDER, "Polla - *.xlsm")))
    datos = []
    for ruta in archivos:
        wb = load_workbook(ruta, read_only=True, data_only=False)
        nombre = wb["CONFIG"]["A2"].value
        clave = wb["CONFIG"]["B2"].value
        ws = wb[nombre]
        apuestas = {}
        for n in range(1, 73):
            gl = ws.cell(row=n + 1, column=6).value
            gv = ws.cell(row=n + 1, column=7).value
            if gl is not None and gv is not None:
                apuestas[n] = (int(gl), int(gv))
        wb.close()
        nombres.append(nombre)
        datos.append((nombre, clave, apuestas))

    db.init_db(nombres)

    # claves -> PIN
    for nombre, clave, _ in datos:
        if clave not in (None, ""):
            with db.engine.begin() as cx:
                from sqlalchemy import text
                cx.execute(text("UPDATE participantes SET pin=:p WHERE nombre=:n"),
                           {"p": db.hash_pin(str(clave)), "n": nombre})

    # apuestas
    total_ap = 0
    for nombre, _, apuestas in datos:
        for n, (gl, gv) in apuestas.items():
            db.set_apuesta(nombre, n, gl, gv)
            total_ap += 1

    # ajustes (correcciones admin) y resultados del TABLERO
    if os.path.exists(TABLERO):
        wb = load_workbook(TABLERO, read_only=True, data_only=False)
        if "AJUSTES" in wb.sheetnames:
            for fila in wb["AJUSTES"].iter_rows(min_row=2, values_only=True):
                nom, m, gl, gv = (list(fila) + [None] * 4)[:4]
                if nom and m is not None and gl is not None and gv is not None:
                    db.set_apuesta(str(nom).strip(), int(m), int(gl), int(gv))
        res = wb["RESULTADOS"]
        nres = 0
        for n in range(1, 73):
            gl = res.cell(row=n + 1, column=7).value
            gv = res.cell(row=n + 1, column=8).value
            if gl is not None and gv is not None:
                db.set_resultado(n, int(gl), int(gv))
                nres += 1
        wb.close()
    else:
        nres = 0

    print(f"Migrados: {len(nombres)} participantes, {total_ap} apuestas, {nres} resultados.")
    print(f"Base: {db.DATABASE_URL}")


if __name__ == "__main__":
    main()
